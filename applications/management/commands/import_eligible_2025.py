from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation

from students.models import EligibleStudent2025


def clean(v):
    return (str(v).strip() if v is not None else "")

def to_decimal(v):
    v = clean(v)
    if not v:
        return None
    try:
        return Decimal(v)
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = "Import 2025 eligible students (excluding final year) from an Excel file."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)

    def handle(self, *args, **opts):
        path = opts["xlsx_path"]
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        # Read header row and map columns (handles trailing spaces like 'First Name ')
        headers = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}

        required = ["First Name", "First Name ", "Surname", "Institution"]
        if not any(h in col for h in ["First Name", "First Name "]):
            raise Exception(f"Missing First Name column. Found headers: {headers}")

        first_name_key = "First Name" if "First Name" in col else "First Name "
        course_key = "Course" if "Course" in col else ("Course " if "Course " in col else None)

        created = 0
        skipped = 0

        # Optional: clear existing imports if you want a clean re-import
        # EligibleStudent2025.objects.all().delete()

        batch = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            first = clean(r[col[first_name_key]])
            last = clean(r[col["Surname"]])
            inst = clean(r[col["Institution"]])

            if not first or not last or not inst:
                skipped += 1
                continue

            obj = EligibleStudent2025(
                first_name=first,
                surname=last,
                gender=clean(r[col.get("Gender", -1)]) if "Gender" in col else "",
                institution=inst,
                course=clean(r[col.get(course_key, -1)]) if course_key and course_key in col else "",
                tuition_fee=to_decimal(r[col.get("Tuition Fee", -1)]) if "Tuition Fee" in col else None,
                district=clean(r[col.get("District", -1)]) if "District" in col else "",
                year_of_study=clean(r[col.get("Year Of Study", -1)]) if "Year Of Study" in col else "",
            )
            batch.append(obj)

            if len(batch) >= 1000:
                EligibleStudent2025.objects.bulk_create(batch, ignore_conflicts=True)
                created += len(batch)
                batch = []

        if batch:
            EligibleStudent2025.objects.bulk_create(batch, ignore_conflicts=True)
            created += len(batch)

        self.stdout.write(self.style.SUCCESS(
            f"Import complete. Inserted ~{created} rows. Skipped {skipped}."
        ))
